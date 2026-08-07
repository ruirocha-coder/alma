# tools/documentos_empresa.py — Documentos e Ficheiros do Basecamp, espalhados
# por vários projetos, como base de conhecimento da empresa para a Alma consultar.
#
# Usa o endpoint global de recordings (o mesmo que tarefas_e_cards_atrasados já
# usa para Todos e Cards) filtrado por type=Document e type=Upload — dá acesso a
# tudo o que a conta da Alma já vê em qualquer projeto, sem ter de percorrer as
# pastas (Vaults) de cada projeto uma a uma.
import io, os, time
from email import policy
from email.parser import BytesParser
from bs4 import BeautifulSoup
from pypdf import PdfReader
from docx import Document as DocxDocument
from openpyxl import load_workbook
from tools import basecamp, visao

_cache = {}  # {"lista": (timestamp, lista)}
TTL = 900  # segundos — documentos de empresa não mudam a cada minuto

# um documento de negócio real (proposta, cadeia de emails) facilmente
# ultrapassa umas páginas — informação importante (ex: condições comerciais)
# muitas vezes só aparece mais à frente, não logo no início, e alguns
# documentos de referência (ex: o "fluxograma" com emails consolidados da
# empresa) têm mesmo de ser lidos por inteiro, não parcialmente. Subido de
# 50000 para 150000 (Rui, 2026-08-06): uma folha de cálculo com uma folha
# por mês (ex: um inventário anual) facilmente passava dos 50000 e perdia
# meses inteiros sem aviso — 150000 cobre isso com folga, mantendo ainda
# um limite para ficheiros verdadeiramente enormes (ver _extrair_xlsx
# para o corte seguro por folha completa, nunca a meio de uma).
LIMITE_CARATERES_DOCUMENTO = 150000

TIPOS_DE_FICHEIRO_LEGIVEIS = {
    "application/pdf",
    "text/plain",
    "text/csv",
    "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "message/rfc822",
}

# quando alguém arrasta um email (.eml) ou uma folha de cálculo (.xlsx)
# para o Basecamp, o browser nem sempre reporta um content_type útil
# (fica "application/octet-stream" ou vazio) — a extensão do ficheiro é o
# sinal mais fiável nesses casos. Só serve de reserva quando o
# content_type não identifica nada por si só.
_EXTENSAO_PARA_TIPO = {
    ".eml": "message/rfc822",
    ".xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
}
_TIPOS_INCONCLUSIVOS = {"", "application/octet-stream", "binary/octet-stream"}

def _tipo_efetivo(ctype: str, filename: str) -> str:
    """O content_type a usar de facto para decidir como extrair o texto —
    normalmente o que o Basecamp devolve, mas com reserva pela extensão do
    nome do ficheiro quando esse content_type não diz nada (ex: um .eml
    identificado só como "application/octet-stream")."""
    ctype = ctype or ""
    if ctype not in _TIPOS_INCONCLUSIVOS:
        return ctype
    ext = os.path.splitext(filename or "")[1].lower()
    return _EXTENSAO_PARA_TIPO.get(ext, ctype)

def _texto_simples(html: str) -> str:
    if not html:
        return ""
    return BeautifulSoup(html, "html.parser").get_text(" ", strip=True)

def _extrair_email(bruto: bytes) -> str:
    """Extrai de um ficheiro de email (.eml) os campos úteis (de/para/
    assunto/data) e o corpo em texto — prefere a versão texto simples do
    email, e limpa o HTML da versão em html quando não há texto simples."""
    msg = BytesParser(policy=policy.default).parsebytes(bruto)
    cabecalho = [
        f"De: {msg.get('from', '(desconhecido)')}",
        f"Para: {msg.get('to', '(desconhecido)')}",
        f"Assunto: {msg.get('subject', '(sem assunto)')}",
        f"Data: {msg.get('date', '(sem data)')}",
    ]
    corpo = msg.get_body(preferencelist=("plain", "html"))
    texto_corpo = ""
    if corpo is not None:
        texto_corpo = corpo.get_content()
        if corpo.get_content_type() == "text/html":
            texto_corpo = _texto_simples(texto_corpo)
    return ("\n".join(cabecalho) + "\n\n" + texto_corpo.strip()).strip()

def _listar_bruto(forcar: bool = False) -> list[dict]:
    """`forcar=True` ignora a cache e vai sempre buscar a lista atual ao
    Basecamp — usado quando quem procura um documento específico não o
    encontra na lista em cache, para não desistir com base numa lista que
    pode estar até 15 minutos desatualizada (ex: documento criado/
    renomeado/partilhado há pouco)."""
    if not forcar and "lista" in _cache:
        ts, lista = _cache["lista"]
        if time.time() - ts < TTL:
            return lista
    itens = []
    for tipo in ("Document", "Upload"):
        registos = basecamp._get_paginado(
            f"{basecamp._base_url()}/projects/recordings.json",
            params={"type": tipo, "status": "active"}, etiqueta=tipo)
        for r in registos:
            itens.append({
                "id": r["id"],
                "tipo": "documento" if tipo == "Document" else "ficheiro",
                "titulo": r.get("title") or r.get("filename") or "(sem título)",
                "projeto": (r.get("bucket") or {}).get("name"),
                "pasta": (r.get("parent") or {}).get("title"),
                "url": r["url"],
                "app_url": r.get("app_url"),
                "content_type": r.get("content_type"),
                "filename": r.get("filename"),
                "download_url": r.get("download_url"),
            })
    _cache["lista"] = (time.time(), itens)
    return itens

def procurar_documentos_empresa(pesquisa: str) -> list[dict]:
    """Procura documentos e ficheiros da empresa no Basecamp (id, título, projeto, pasta),
    em todos os projetos onde a Alma tem acesso. Filtra por título, projeto ou pasta
    conterem o termo indicado — há mais de mil documentos no total, por isso listar tudo
    de uma vez não é prático; usa um termo relacionado com o que procuras (ex: nome do
    documento, do projeto ou do tema). Devolve no máximo 40 resultados."""
    termo = pesquisa.lower().strip()
    correspondem = [
        item for item in _listar_bruto()
        if termo in item["titulo"].lower()
        or termo in (item.get("projeto") or "").lower()
        or termo in (item.get("pasta") or "").lower()
    ]
    return [{k: v for k, v in item.items() if k in ("id", "tipo", "titulo", "projeto", "pasta")}
            for item in correspondem[:40]]

def _texto_folha_xlsx(livro, nome_folha: str) -> str:
    linhas_texto = []
    for linha in livro[nome_folha].iter_rows(values_only=True):
        celulas = ["" if v is None else str(v) for v in linha]
        if any(c.strip() for c in celulas):
            linhas_texto.append(" | ".join(celulas))
    return "\n".join(linhas_texto)

def _extrair_xlsx(bruto: bytes, folha: str = None, limite_caracteres: int = None) -> str:
    """Converte uma folha de cálculo (.xlsx) em texto tabular simples — usa
    valores já calculados (data_only=True), nunca fórmulas em bruto, para
    a Alma ler o mesmo que uma pessoa vê ao abrir o ficheiro. Ignora
    linhas completamente vazias.

    Com `folha` indicado, devolve só essa folha (por nome exato ou parcial,
    sem sensibilidade a maiúsculas) — usa isto para ficheiros grandes em
    vez de pedir tudo de uma vez (ver ler_folha_excel_anexo/
    ler_documento_empresa). Sem `folha`, devolve todas as folhas com
    conteúdo, uma secção por folha; se isso ultrapassar
    `limite_caracteres`, PÁRA sempre no limite de uma folha completa —
    nunca corta uma folha a meio — e acrescenta uma nota clara com os
    nomes das folhas que ficaram de fora, para nunca perder dados em
    silêncio. Bug real, 2026-08-06: um inventário com 12 folhas (uma por
    mês) foi cortado a meio da folha de março, perdendo junho por
    completo, sem nenhum aviso disso na resposta."""
    livro = load_workbook(io.BytesIO(bruto), data_only=True, read_only=True)
    nomes_folhas = livro.sheetnames

    if folha is not None:
        correspondente = next((n for n in nomes_folhas if n.lower() == folha.lower()), None)
        if correspondente is None:
            correspondente = next((n for n in nomes_folhas if folha.lower() in n.lower()), None)
        if correspondente is None:
            return f"(não encontrei a folha {folha!r} — folhas disponíveis: {', '.join(nomes_folhas)})"
        return f"[Folha: {correspondente}]\n{_texto_folha_xlsx(livro, correspondente)}"

    blocos, folhas_omitidas, total = [], [], 0
    for nome_folha in nomes_folhas:
        texto_folha = _texto_folha_xlsx(livro, nome_folha)
        if not texto_folha:
            continue
        bloco = f"[Folha: {nome_folha}]\n{texto_folha}"
        if limite_caracteres is not None and blocos and total + len(bloco) > limite_caracteres:
            folhas_omitidas.append(nome_folha)
            continue
        blocos.append(bloco)
        total += len(bloco)
    texto_final = "\n\n".join(blocos).strip()
    if folhas_omitidas:
        texto_final += (f"\n\n(nota: ficheiro extenso — {len(folhas_omitidas)} folha(s) não incluída(s) "
                        f"aqui: {', '.join(folhas_omitidas)}. Pede uma destas pelo nome para a leres na "
                        f"íntegra, em vez de assumires que não existe ou está vazia.)")
    return texto_final

def _extrair_por_tipo(bruto: bytes, ctype: str, folha: str = None) -> str:
    """Extrai texto de bytes crus dado o content_type — partilhado entre
    ficheiros (Uploads) e anexos embutidos dentro de Documentos nativos do
    Basecamp (ver _ler_conteudo). `folha` só se aplica a folhas de cálculo
    (.xlsx) — ignorado para os outros tipos."""
    if ctype in visao.TIPOS_DE_IMAGEM:
        return visao.descrever_imagem(bruto, ctype)
    if ctype == "application/pdf":
        leitor = PdfReader(io.BytesIO(bruto))
        texto = "\n".join(pagina.extract_text() or "" for pagina in leitor.pages).strip()
        if not texto:
            # sem texto extraível — provavelmente um PDF só de design/imagem/
            # scan; descreve página a página em vez de só a primeira, para
            # não perder conteúdo (ex: um contrato/proposta escaneado).
            try:
                texto = visao.descrever_pdf_escaneado(bruto)
            except Exception as e:
                texto = f"(não consegui extrair texto nem imagem deste PDF: {e})"
        return texto
    if ctype == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
        doc = DocxDocument(io.BytesIO(bruto))
        return "\n".join(paragrafo.text for paragrafo in doc.paragraphs).strip()
    if ctype == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        return _extrair_xlsx(bruto, folha=folha, limite_caracteres=LIMITE_CARATERES_DOCUMENTO)
    if ctype in ("text/plain", "text/csv"):
        return bruto.decode("utf-8", errors="ignore")
    if ctype == "message/rfc822":
        return _extrair_email(bruto)
    return None

def _ler_conteudo(item: dict, folha: str = None) -> str:
    """Extrai o texto de um documento/ficheiro (item de _listar_bruto()) —
    partilhado entre ler_documento_empresa (um id específico) e outros
    sítios que precisem de ler vários documentos de uma vez (ex: um projeto
    inteiro tratado como fonte de confiança). Devolve None quando o tipo de
    ficheiro não é legível (quem chamar decide o que fazer nesse caso).
    `folha` só se aplica quando o ficheiro é uma folha de cálculo (.xlsx)."""
    if item["tipo"] == "documento":
        # um Documento nativo do Basecamp pode ter o texto todo escrito nele
        # próprio, ou pode ser só um invólucro com o conteúdo real num
        # ficheiro anexado lá dentro (PDF, Word, imagem) — lê os dois e
        # junta, em vez de assumir que é sempre um ou outro.
        completo = basecamp.obter_recording(item["url"])
        texto_wrapper = _texto_simples(completo.get("content", "")).strip()
        partes = [texto_wrapper] if texto_wrapper else []
        for anexo in completo.get("content_attachments") or []:
            ctype_anexo = _tipo_efetivo(anexo.get("content_type"), anexo.get("filename"))
            try:
                bruto_anexo = basecamp._get_bytes(anexo["download_url"])
                texto_anexo = _extrair_por_tipo(bruto_anexo, ctype_anexo, folha=folha)
                if texto_anexo:
                    partes.append(texto_anexo)
            except Exception as e:
                partes.append(f"(erro ao ler um anexo deste documento: {e})")
        texto_final = "\n\n".join(partes).strip()
        # o corte aqui é só uma rede de segurança adicional — a extração de
        # cada anexo (ver _extrair_xlsx) já para em limites de folha
        # completa para uma folha de cálculo, nunca a meio.
        return texto_final[:LIMITE_CARATERES_DOCUMENTO] if texto_final else None

    ctype = _tipo_efetivo(item.get("content_type"), item.get("filename"))

    if ctype in visao.TIPOS_DE_IMAGEM:
        bruto = basecamp._get_bytes(item["download_url"])
        return visao.descrever_imagem(bruto, ctype)

    if ctype not in TIPOS_DE_FICHEIRO_LEGIVEIS:
        return None

    bruto = basecamp._get_bytes(item["download_url"])
    texto = _extrair_por_tipo(bruto, ctype, folha=folha)
    return texto[:LIMITE_CARATERES_DOCUMENTO] if texto else texto

def ler_documento_empresa(id: int, folha: str = None) -> dict:
    """Lê o conteúdo de texto de um documento ou ficheiro da empresa, pelo id
    (de listar_documentos_empresa). Suporta documentos nativos do Basecamp,
    PDF, Word (.docx), Excel (.xlsx — cada folha de cálculo é convertida em
    texto tabular, com valores já calculados, nunca fórmulas em bruto),
    email (.eml), texto simples e CSV.

    Se o ficheiro for uma folha de cálculo (.xlsx) grande (várias folhas,
    ex: uma por mês), a primeira leitura (sem `folha`) devolve todas as
    folhas que couberem, e diz claramente quais ficaram de fora, se
    alguma — nunca corta uma folha a meio. Volta a chamar esta função com
    `folha` (o nome, ou parte dele, ex: "Junho") para leres essa folha na
    íntegra, em vez de assumires que não existe."""
    item = next((i for i in _listar_bruto() if i["id"] == id), None)
    if not item:
        return {"erro": "documento não encontrado — confirma o id com procurar_documentos_empresa"}

    conteudo = _ler_conteudo(item, folha=folha)
    if conteudo is None:
        if item["tipo"] == "documento":
            return {"erro": "este documento parece estar vazio (sem texto e sem anexos legíveis)",
                    "titulo": item["titulo"], "app_url": item.get("app_url")}
        ctype = item.get("content_type") or ""
        return {"erro": f"não consigo ler o conteúdo deste tipo de ficheiro ({ctype or item.get('filename')})",
                "titulo": item["titulo"], "app_url": item.get("app_url")}
    return {"titulo": item["titulo"], "conteudo": conteudo}

def ler_anexos_registo_basecamp(url: str) -> dict:
    """Lê o conteúdo dos ficheiros anexados diretamente a um registo do
    Basecamp — a descrição de uma tarefa/card, OU um comentário (ex: um PDF
    de desenho técnico ou especificações de um produto partilhado num
    comentário, não só na tarefa em si). `url` é o url da própria API desse
    registo (o campo "url" que já vem no contexto ou em ler_comentarios —
    nunca inventes ou reconstruas este url a partir só do id: o Basecamp
    aninha os recordings sob o bucket do projeto, um formato tipo
    ".../recordings/{id}.json" na raiz não existe e dá sempre 404). Não é
    para documentos/ficheiros avulsos — para isso usa
    procurar_documentos_empresa/ler_documento_empresa."""
    try:
        recording = basecamp.obter_recording(url)
    except Exception as e:
        return {"erro": f"não consegui aceder a este registo do Basecamp: {e}"}

    # bug real confirmado ao vivo (Beatriz, 2026-07-27, contra a API real do
    # Basecamp): o array de anexos embutidos chama-se "content_attachments"
    # só para tipos com um campo "content" (ex: Message) — uma tarefa/card do
    # Card Table tem as notas no campo "description", e os anexos embutidos
    # nela vêm em "description_attachments", um nome diferente. Esta função
    # só olhava para "content_attachments", por isso NUNCA encontrava PDFs
    # anexados diretamente às notas de um card (dizia sempre "não tem
    # ficheiros anexados", mesmo havendo — confirmado com o card real
    # "Ana Fraião", id 9577718481, com "OR 2026_13.pdf" e outro PDF
    # anexados diretamente na descrição). Junta os dois, nunca só um.
    anexos = (recording.get("content_attachments") or []) + (recording.get("description_attachments") or [])
    if not anexos:
        return {"anexos": [], "aviso": "este registo não tem ficheiros anexados diretamente"}

    resultados = []
    for anexo in anexos:
        nome = anexo.get("filename") or anexo.get("name") or "(sem nome)"
        ctype = _tipo_efetivo(anexo.get("content_type"), nome)
        try:
            bruto = basecamp._get_bytes(anexo["download_url"])
            texto = _extrair_por_tipo(bruto, ctype)
            resultados.append({"ficheiro": nome, "conteudo": (texto or "(sem texto legível)")[:LIMITE_CARATERES_DOCUMENTO]})
        except Exception as e:
            resultados.append({"ficheiro": nome, "erro": str(e)})
    return {"anexos": resultados}

def ler_folha_excel_anexo(url: str, ficheiro: str, folha: str = None) -> dict:
    """Lê uma folha específica (ou lista as folhas disponíveis, se `folha`
    não for indicado) de um ficheiro Excel (.xlsx) anexado diretamente a
    um registo do Basecamp — pensado para ficheiros com várias folhas
    (ex: um inventário com uma folha por mês) que ler_anexos_registo_basecamp
    pode não conseguir trazer inteiros de uma vez. `url` é o url da
    própria API do registo (tarefa/card ou comentário — nunca inventado);
    `ficheiro` é o nome (ou parte do nome) do anexo, tal como aparece em
    ler_anexos_registo_basecamp."""
    try:
        recording = basecamp.obter_recording(url)
    except Exception as e:
        return {"erro": f"não consegui aceder a este registo do Basecamp: {e}"}

    anexos = (recording.get("content_attachments") or []) + (recording.get("description_attachments") or [])
    termo = ficheiro.lower()
    anexo = next((a for a in anexos if termo in (a.get("filename") or a.get("name") or "").lower()), None)
    if anexo is None:
        nomes = [a.get("filename") or a.get("name") or "(sem nome)" for a in anexos]
        return {"erro": f"não encontrei um anexo com esse nome neste registo — anexos disponíveis: {nomes}"}

    nome_real = anexo.get("filename") or anexo.get("name") or ""
    if not nome_real.lower().endswith(".xlsx"):
        return {"erro": f"{nome_real!r} não é um ficheiro .xlsx — usa ler_anexos_registo_basecamp para este anexo"}

    try:
        bruto = basecamp._get_bytes(anexo["download_url"])
    except Exception as e:
        return {"erro": f"não consegui descarregar o ficheiro: {e}"}

    if folha is None:
        livro = load_workbook(io.BytesIO(bruto), data_only=True, read_only=True)
        return {"ficheiro": nome_real, "folhas_disponiveis": livro.sheetnames}

    conteudo = _extrair_xlsx(bruto, folha=folha)
    return {"ficheiro": nome_real, "conteudo": conteudo}

TOOLS_DOCUMENTOS_EMPRESA = [
    {
        "name": "procurar_documentos_empresa",
        "description": "Procura documentos e ficheiros da empresa guardados no Basecamp, em todos os projetos (id, tipo, título, projeto, pasta), por um termo no título/projeto/pasta. Usa isto para descobrir que documentos existem antes de leres um com ler_documento_empresa.",
        "input_schema": {
            "type": "object",
            "properties": {"pesquisa": {"type": "string"}},
            "required": ["pesquisa"]
        }
    },
    {
        "name": "ler_documento_empresa",
        "description": "Lê o conteúdo de texto de um documento ou ficheiro da empresa, pelo id devolvido por procurar_documentos_empresa. Suporta documentos nativos do Basecamp, PDF (mesmo quando o PDF é só design/imagem sem texto), Word (.docx), Excel (.xlsx — cada folha vem como texto tabular, com os valores já calculados, nunca fórmulas em bruto), email (.eml — lê de/para/assunto/data e o corpo do email), imagens (JPG, PNG, GIF, WebP — descritas/transcritas por visão), texto simples e CSV — outros formatos devolvem um erro com o link para abrir manualmente. Se for uma folha de cálculo grande (várias folhas, ex: uma por mês), a resposta diz claramente quais folhas ficaram de fora, se alguma — chama outra vez com `folha` para leres essa em concreto, nunca assumas que não existe.",
        "input_schema": {
            "type": "object",
            "properties": {
                "id": {"type": "integer"},
                "folha": {"type": "string", "description": "só para .xlsx: nome (ou parte do nome) de uma folha específica a ler, ex: \"Junho\" — omite para ler todas as folhas que couberem"}
            },
            "required": ["id"]
        }
    },
    {
        "name": "ler_anexos_registo_basecamp",
        "description": "Lê o conteúdo dos ficheiros anexados diretamente a um registo do Basecamp — a descrição de uma tarefa/card, OU um comentário específico (ex: um PDF de desenho técnico, uma fatura ou um orçamento com os produtos de uma encomenda, suporta os mesmos formatos que ler_documento_empresa). Usa isto quando a pergunta precisar de informação que só está nesses anexos (ex: \"qual o tamanho da prateleira?\", medidas, especificações, que produtos estão nesta encomenda para prever o tempo de montagem, ou um pedido para resumir contas/valores quando há faturas, recibos ou comprovativos anexados) — não leias por rotina em toda tarefa/card, só quando a pergunta for mesmo sobre isso. `url` é sempre o url da própria API desse registo — o campo \"url_api\" de um card devolvido por procurar_cards_basecamp/estado_projeto_basecamp, o \"url\" que já vem no contexto, ou o de um comentário específico devolvido por ler_comentarios — nunca inventes este url a partir só de um número/id, o Basecamp aninha os recordings sob o bucket do projeto e um url reconstruído à mão dá sempre 404. IMPORTANTE: se o anexo lido for uma fatura/orçamento com o seu próprio campo de morada, essa é a morada fiscal/de faturação do cliente — nunca a uses para responder sobre a morada de ENTREGA (essa vem sempre só das notas do card, nunca de um anexo). Se um anexo for uma folha de cálculo (.xlsx) grande (várias folhas, ex: uma por mês), esta função pode não trazer todas as folhas de uma vez — usa então ler_folha_excel_anexo com o mesmo `url` para leres uma folha específica na íntegra.",
        "input_schema": {
            "type": "object",
            "properties": {"url": {"type": "string", "description": "o url da própria API da tarefa/card ou do comentário — nunca inventado"}},
            "required": ["url"]
        }
    },
    {
        "name": "ler_folha_excel_anexo",
        "description": "Lê uma folha específica (ou lista as folhas disponíveis, se `folha` não for indicado) de um ficheiro Excel (.xlsx) anexado diretamente a um registo do Basecamp (tarefa/card ou comentário) — usa isto para folhas de cálculo grandes (várias folhas/meses), em vez de ler_anexos_registo_basecamp, que pode não conseguir trazer o ficheiro inteiro de uma vez. Primeiro chama sem `folha` para veres os nomes disponíveis, depois chama outra vez com o nome exato (ou parcial, ex: \"Junho\") para leres essa folha completa.",
        "input_schema": {
            "type": "object",
            "properties": {
                "url": {"type": "string", "description": "o url da própria API da tarefa/card ou do comentário onde o ficheiro está anexado — nunca inventado"},
                "ficheiro": {"type": "string", "description": "nome (ou parte do nome) do ficheiro .xlsx, tal como aparece em ler_anexos_registo_basecamp"},
                "folha": {"type": "string", "description": "nome (ou parte do nome) da folha a ler — omite para listares só os nomes das folhas disponíveis"}
            },
            "required": ["url", "ficheiro"]
        }
    }
]
