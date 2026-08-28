# tools/ficheiros.py — extrai texto de um ficheiro em bruto (bytes), qualquer
# que seja a origem (aqui: anexos enviados na consola de chat). Reaproveita a
# mesma lógica de leitura já usada para documentos do Basecamp.
import io, os
from pypdf import PdfReader
from docx import Document as DocxDocument
from openpyxl import load_workbook
from tools import visao

TIPOS_DE_TEXTO = {"text/plain", "text/csv", "text/markdown"}
EXTENSOES_DE_TEXTO = (".txt", ".csv", ".md")

# ver tools/documentos_empresa.LIMITE_CARATERES_DOCUMENTO — mesma razão e
# mesmo valor: uma folha de cálculo com uma folha por mês facilmente passa
# de limites mais baixos, perdendo meses inteiros sem aviso.
LIMITE_CARATERES_XLSX = 150000

# o browser nem sempre reporta um content_type de imagem fiável para um
# upload (fica genérico ou vazio, dependendo do browser/sistema) — a
# extensão do nome do ficheiro é o sinal de reserva, tal como já se faz
# para anexos do Basecamp (ver tools/documentos_empresa._tipo_efetivo).
_EXTENSAO_PARA_TIPO_IMAGEM = {
    ".jpg": "image/jpeg", ".jpeg": "image/jpeg",
    ".png": "image/png", ".gif": "image/gif", ".webp": "image/webp",
}

def extrair_texto(bruto: bytes, content_type: str, filename: str = "") -> str | None:
    """Devolve o texto extraído do ficheiro, ou None se o tipo não for suportado."""
    content_type = (content_type or "").split(";")[0].strip().lower()
    nome = (filename or "").lower()
    extensao = os.path.splitext(nome)[1]

    if content_type in visao.TIPOS_DE_IMAGEM or extensao in _EXTENSAO_PARA_TIPO_IMAGEM:
        media_type = content_type if content_type in visao.TIPOS_DE_IMAGEM else _EXTENSAO_PARA_TIPO_IMAGEM[extensao]
        return visao.descrever_imagem(bruto, media_type)

    if content_type == "application/pdf" or nome.endswith(".pdf"):
        leitor = PdfReader(io.BytesIO(bruto))
        texto = "\n".join(pagina.extract_text() or "" for pagina in leitor.pages).strip()
        if not texto:
            # sem texto extraível — provavelmente um PDF só de design/imagem/
            # scan; descreve página a página em vez de só a primeira
            try:
                return visao.descrever_pdf_escaneado(bruto)
            except Exception as e:
                return f"(não consegui extrair texto nem imagem deste PDF: {e})"
        try:
            # um PDF pode ter texto (códigos de cor, anotações) E imagens/
            # desenhos vetoriais (o símbolo/tipografia de um logótipo) — só
            # devolver o texto perdia esse conteúdo visual em silêncio
            # (bug real, 2026-08-28: logo enviado em PDF só transmitiu os
            # códigos de cor, nunca o desenho do logótipo em si)
            if visao.pdf_tem_conteudo_visual(bruto):
                return f"{texto}\n\n[Conteúdo visual do PDF]\n{visao.descrever_pdf_escaneado(bruto)}"
        except Exception:
            pass
        return texto

    if (content_type == "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
            or nome.endswith(".docx")):
        doc = DocxDocument(io.BytesIO(bruto))
        return "\n".join(paragrafo.text for paragrafo in doc.paragraphs).strip()

    if (content_type == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            or nome.endswith(".xlsx")):
        livro = load_workbook(io.BytesIO(bruto), data_only=True, read_only=True)

        def _texto_folha(nome_folha):
            linhas_texto = []
            for linha in livro[nome_folha].iter_rows(values_only=True):
                celulas = ["" if v is None else str(v) for v in linha]
                if any(c.strip() for c in celulas):
                    linhas_texto.append(" | ".join(celulas))
            return "\n".join(linhas_texto)

        # nunca corta uma folha a meio — pára sempre num limite de folha
        # completa, e diz claramente quais ficaram de fora, se alguma (bug
        # real, 2026-08-06: um inventário com 12 folhas foi cortado a meio,
        # perdendo meses inteiros sem nenhum aviso na resposta).
        blocos, folhas_omitidas, total = [], [], 0
        for nome_folha in livro.sheetnames:
            texto_folha = _texto_folha(nome_folha)
            if not texto_folha:
                continue
            bloco = f"[Folha: {nome_folha}]\n{texto_folha}"
            if blocos and total + len(bloco) > LIMITE_CARATERES_XLSX:
                folhas_omitidas.append(nome_folha)
                continue
            blocos.append(bloco)
            total += len(bloco)
        texto_final = "\n\n".join(blocos).strip()
        if folhas_omitidas:
            texto_final += (f"\n\n(nota: ficheiro extenso — {len(folhas_omitidas)} folha(s) não incluída(s) "
                            f"aqui: {', '.join(folhas_omitidas)}. Pede para reenviar só essa folha em separado "
                            "se precisares dela.)")
        return texto_final

    if content_type in TIPOS_DE_TEXTO or nome.endswith(EXTENSOES_DE_TEXTO):
        return bruto.decode("utf-8", errors="ignore")

    return None
