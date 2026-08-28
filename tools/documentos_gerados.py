# tools/documentos_gerados.py — gera documentos formatados (PDF ou Excel)
# para partilhar na conversa (ex: um relatório longo, uma proposta, uma
# folha de cálculo de dados) em vez de despejar tudo como texto corrido no
# chat. Guardado em Postgres, não em disco — o Railway não persiste
# ficheiros locais entre deploys — e servido pelo próprio endpoint de
# download em main.py.
import os
import json
import markdown as _markdown
from xhtml2pdf import pisa
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
import io
import db

_CSS = """
<style>
  @page { size: A4; margin: 2.2cm; }
  body { font-family: Helvetica, Arial, sans-serif; font-size: 11pt; line-height: 1.5; color: #1a1a1a; }
  h1 { font-size: 20pt; margin-top: 0; }
  h2 { font-size: 15pt; margin-top: 1.2em; }
  h3 { font-size: 12.5pt; margin-top: 1em; }
  table { border-collapse: collapse; width: 100%; margin: 0.8em 0; }
  th, td { border: 1px solid #999; padding: 6px 8px; text-align: left; }
  code { font-family: Courier, monospace; background: #f0f0f0; padding: 1px 4px; }
  pre { background: #f0f0f0; padding: 8px; }
</style>
"""

def _markdown_para_html(titulo: str, conteudo_markdown: str) -> str:
    corpo = _markdown.markdown(conteudo_markdown, extensions=["extra", "sane_lists"])
    return f"<html><head>{_CSS}</head><body><h1>{titulo}</h1>{corpo}</body></html>"

def gerar_pdf(utilizador: str, titulo: str, conteudo_markdown: str) -> dict:
    """Gera um documento PDF formatado a partir de conteúdo em markdown
    (títulos, negrito, listas, tabelas — a mesma sintaxe que já usas nas
    respostas normais) e devolve um url para o partilhares na conversa.
    Usa isto sempre que o pedido for um documento longo/formal (um
    relatório, uma proposta, um resumo estruturado de várias páginas), ou
    sempre que pedirem explicitamente um PDF — em vez de escreveres tudo
    como texto corrido no chat. Inclui sempre o url devolvido na tua
    resposta em formato de link markdown (ex: "[título](url)"), para a
    pessoa poder abrir/descarregar o documento.

    Guarda também o markdown-fonte (não só o PDF já compilado), associado
    a quem pediu — pedido do Rui (2026-08-03): sem isto, pedir para
    reaproveitar/converter um documento já feito ("passa o último PDF a
    Excel") não tinha a que voltar, só o PDF final compilado. Ver
    db.documentos_gerados_recentes (injetado automaticamente no contexto)
    e obter_conteudo_documento_gerado."""
    html = _markdown_para_html(titulo, conteudo_markdown)
    buffer = io.BytesIO()
    resultado = pisa.CreatePDF(html, dest=buffer)
    if resultado.err:
        return {"erro": "não consegui gerar o PDF a partir deste conteúdo"}
    id_gerado = db.guardar_documento_gerado(utilizador, titulo, buffer.getvalue(), conteudo_markdown, formato="pdf")
    url = f"{os.environ['ALMA_APP_URL'].rstrip('/')}/documentos-gerados/{id_gerado}"
    return {"titulo": titulo, "url": url}

def gerar_excel(utilizador: str, titulo: str, colunas: list, linhas: list,
                subtitulo: str = None, linhas_destacadas: list = None) -> dict:
    """Gera uma folha de cálculo Excel (.xlsx) REAL, pronta a descarregar, a
    partir de colunas e linhas de dados, e devolve um url para a
    partilhares na conversa. Usa isto sempre que pedirem dados em Excel/
    folha de cálculo, ou para converter uma tabela/lista/documento já
    feito para esse formato. NUNCA escrevas os dados como texto/tabela a
    fingir que geraste um Excel — sem chamar esta função não existe
    ficheiro nenhum para a pessoa descarregar. Inclui sempre o url
    devolvido na tua resposta em formato de link markdown (ex:
    "[título](url)"), para a pessoa poder abrir/descarregar o documento.

    Guarda também colunas+linhas em bruto (não só o .xlsx já compilado),
    para poderes reler/reaproveitar depois — ver
    obter_conteudo_documento_gerado, mesmo princípio de gerar_pdf."""
    ws_titulo = (titulo or "Folha1")[:31]  # limite do Excel para nomes de folha
    wb = Workbook()
    ws = wb.active
    ws.title = ws_titulo

    linha_atual = 1
    ws.cell(row=linha_atual, column=1, value=titulo).font = Font(bold=True, size=14)
    if subtitulo:
        linha_atual += 1
        ws.cell(row=linha_atual, column=1, value=subtitulo).font = Font(italic=True, color="555555")
    linha_atual += 2  # linha em branco antes do cabeçalho

    linha_cabecalho = linha_atual
    for i, coluna in enumerate(colunas, start=1):
        celula = ws.cell(row=linha_cabecalho, column=i, value=coluna)
        celula.font = Font(bold=True)
        celula.fill = PatternFill("solid", fgColor="E0E0E0")

    # linhas_destacadas: índices (a partir de 0, sobre `linhas`) que a Alma
    # pode marcar visualmente — ex: por preencher, pendentes, a rever —
    # sem precisar de embutir isso no valor das células
    destacadas = set(linhas_destacadas or [])
    for indice, linha in enumerate(linhas):
        linha_folha = linha_cabecalho + 1 + indice
        for i, valor in enumerate(linha, start=1):
            celula = ws.cell(row=linha_folha, column=i, value=valor)
            if indice in destacadas:
                celula.fill = PatternFill("solid", fgColor="FFF2CC")

    # painel congelado logo abaixo do cabeçalho, para ele ficar sempre
    # visível a percorrer muitas linhas
    ws.freeze_panes = ws.cell(row=linha_cabecalho + 1, column=1).coordinate

    for i, coluna in enumerate(colunas, start=1):
        maior = len(str(coluna))
        for linha in linhas:
            if i - 1 < len(linha):
                maior = max(maior, len(str(linha[i - 1])))
        ws.column_dimensions[get_column_letter(i)].width = min(maior + 2, 40)

    buffer = io.BytesIO()
    wb.save(buffer)
    conteudo_fonte = json.dumps(
        {"colunas": colunas, "linhas": linhas, "subtitulo": subtitulo, "linhas_destacadas": linhas_destacadas},
        ensure_ascii=False
    )
    id_gerado = db.guardar_documento_gerado(utilizador, titulo, buffer.getvalue(), conteudo_fonte, formato="xlsx")
    url = f"{os.environ['ALMA_APP_URL'].rstrip('/')}/documentos-gerados/{id_gerado}"
    return {"titulo": titulo, "url": url}

def obter_conteudo_documento_gerado(utilizador: str, id: int) -> dict:
    """Relê o markdown-fonte de um documento já gerado (ver gerar_pdf) —
    usa isto quando pedirem para reaproveitar, atualizar, resumir ou
    converter para outro formato um documento já feito, em vez de dizeres
    que não tens acesso a esses dados. O id pode vir da lista de
    "Documentos que já geraste para esta pessoa" no teu contexto, ou de um
    url /documentos-gerados/{id} partilhado na conversa — não é preciso
    teres sido tu a gerar esse documento nem que seja desta sessão: os
    documentos gerados não são privados por utilizador (a própria rota
    pública que os serve não tem essa restrição)."""
    return db.obter_conteudo_documento_gerado(utilizador, id)

TOOLS_DOCUMENTOS_GERADOS = [
    {
        "name": "gerar_pdf",
        "description": "Gera um documento PDF formatado a partir de conteúdo em markdown (títulos, negrito, listas, tabelas) e devolve um url para partilhares na conversa. Usa isto sempre que o pedido for um documento longo/formal (relatório, proposta, resumo de várias páginas) ou sempre que pedirem explicitamente um PDF, em vez de escreveres tudo como texto corrido no chat. Inclui sempre o url devolvido na tua resposta em formato de link markdown, ex: \"[título](url)\".",
        "input_schema": {
            "type": "object",
            "properties": {
                "titulo": {"type": "string", "description": "título do documento"},
                "conteudo_markdown": {"type": "string", "description": "o conteúdo completo do documento, em markdown (títulos com #, negrito, listas, tabelas)"}
            },
            "required": ["titulo", "conteudo_markdown"]
        }
    },
    {
        "name": "gerar_excel",
        "description": "Gera uma folha de cálculo Excel (.xlsx) REAL, pronta a descarregar, a partir de colunas e linhas de dados, e devolve um url para partilhares na conversa. Usa isto sempre que pedirem dados em Excel/folha de cálculo, ou para converter uma tabela/lista/documento já feito para esse formato. NUNCA escrevas os dados como texto/tabela markdown a fingir que geraste um Excel — sem chamar esta função não existe ficheiro nenhum para a pessoa descarregar. Inclui sempre o url devolvido na tua resposta em formato de link markdown, ex: \"[título](url)\".",
        "input_schema": {
            "type": "object",
            "properties": {
                "titulo": {"type": "string", "description": "título do documento"},
                "subtitulo": {"type": "string", "description": "linha opcional abaixo do título (ex: um total ou resumo)"},
                "colunas": {"type": "array", "items": {"type": "string"}, "description": "cabeçalhos das colunas, por ordem"},
                "linhas": {
                    "type": "array",
                    "items": {"type": "array", "items": {}},
                    "description": "cada linha é uma lista de valores, na mesma ordem das colunas"
                },
                "linhas_destacadas": {
                    "type": "array", "items": {"type": "integer"},
                    "description": "índices (a partir de 0) de linhas de `linhas` a destacar visualmente (ex: por preencher, pendentes) — opcional"
                }
            },
            "required": ["titulo", "colunas", "linhas"]
        }
    },
    {
        "name": "obter_conteudo_documento_gerado",
        "description": "Relê o markdown-fonte de um documento já gerado (ver gerar_pdf), pelo id listado no teu contexto (\"Documentos que já geraste para esta pessoa\") ou pelo id num url /documentos-gerados/{id} partilhado na conversa — funciona para qualquer documento gerado, não só os desta sessão ou desta pessoa. Usa isto sempre que pedirem para reaproveitar, atualizar, resumir ou converter para outro formato (ex: Excel) um documento já feito, em vez de dizeres que não tens essa informação.",
        "input_schema": {
            "type": "object",
            "properties": {
                "id": {"type": "integer", "description": "id do documento, como listado no teu contexto"}
            },
            "required": ["id"]
        }
    }
]
